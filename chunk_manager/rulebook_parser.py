import openpyxl
import json
from pathlib import Path
from typing import Optional

from utils.settings_manager import get_setting

def parse_rulebook_excel(file_path: Path) -> Optional[Path]:
    """
    Parses an Excel workbook defining the topic sentiment distribution for a text corpuse
    with additional parameters. This Excel workbook needs to follow the structure of the
    template workbook.
    
    The workbook can be formatted in one of two ways based on the collection_mode parameter:
    - "word": Defines exact word count ranges for each topic-sentiment pair.
    - "chunk": Defines number of chunks for each topic-sentiment pair.

    This function loads an Excel workbook from the specified path and extracts:
      - 'review_item': from cell A1
      - 'collection_mode': either 'word' or 'chunk' from cell E1
      - 'total': from cell G1 (must be an int > 0)
        - either defines total word count or number of chunks depending on collection_mode
      - 'content_rules': a dictionary mapping topic names to a dictionary containing:
          - 'total_proportion': float from col B
          - 'sentiment_proportion': a tuple of three floats from cols C, D, and E (which must sum to 1)
          - 'chunk_min_wc': int from col H (must be > 0)
          - 'chunk_max_wc': int from col I (must be > chunk_min_wc)
          - 'chunk_pref': mapped from col J if present
            - Represents the preferred number of chunks for this topic-sentiment pair or word count 
              size of each chunk depending on collection_mode
          - 'chunk_wc_distribution': mapped from col K if present
      - 'collection_ranges': a list of dictionaries representing ranges (either word count or chunk count),
          each with keys:
              - 'range': a tuple (start, end) where end > start
              - 'target_fraction': fraction for this range (must sum to 1)

    Args:
        rulebook_name (str): The file name of the Excel workbook, located in the "rulebooks" directory.

    Returns:
        Optional[Path]: Full path to the new JSON file on success.
        None if a validation error occurs.
    """
    
    try:
        # Load the Excel workbook from the rulebooks directory
        if not file_path.exists():
            print(f"parse_rulebook_excel: File does not exist: {file_path}")
            return None
        if file_path.suffix.lower() not in {'.xlsx', '.xlsm'}:
            print(f"parse_rulebook_excel: File is not an Excel workbook: {file_path}")
            return None

        # Load the workbook and ensure the "MAIN" sheet is present
        wb = openpyxl.load_workbook(file_path)
        if "MAIN" not in wb.sheetnames:
            print(f"parse_rulebook_excel: 'MAIN' sheet not found in workbook: {file_path}")
            return None
        ws = wb["MAIN"]

        # Retrieve review_item from cell A1 and validate it
        review_item = ws['A1'].value
        if not isinstance(review_item, str):
            print("parse_rulebook_excel: Invalid value in cell A1 for review_item.")
            return None
        
        # Retrieve collection_mode from cell E1 and validate it
        collection_mode = ws['E1'].value
        if collection_mode not in ('word', 'chunk'):
            print("parse_rulebook_excel: Invalid value in cell E1 for collection_mode.")
            return None

        # Retrieve total from cell G1 and validate it
        total = ws['G1'].value
        if not isinstance(total, int) or total <= 0:
            print("parse_rulebook_excel: Invalid value in cell G1 for total.")
            return None

        content_rules = {}
        # Iterate over rows starting at row 5; include columns A-E and H-K.
        for row_num, row in enumerate(ws.iter_rows(min_row=5, max_col=11, values_only=True), start=5):
            # Column A: topic (index 0)
            topic = row[0]
            # Column B: total_proportion (index 1)
            total_proportion = row[1]
            # Columns C-E: sentiment proportions (indices 2, 3, 4)
            c = row[2]
            d = row[3]
            e = row[4]
            # Column H: chunk_min_wc (index 7)
            chunk_min_wc = row[7]
            # Column I: chunk_max_wc (index 8)
            chunk_max_wc = row[8]
            # Columns J: chunk_count_pref (index 9)
            chunk_pref_raw = row[9]
            # Columns K: chunk_wc_distribution (index 10)
            chunk_wc_distribution_raw = row[10]

            # Terminate iteration if the topic cell is empty
            if topic is None:
                break
            if not isinstance(topic, str):
                print(f"parse_rulebook_excel: Non-string value for topic at row {row_num}.")
                return None

            # Validate total_proportion
            if not isinstance(total_proportion, (int, float)):
                print(f"parse_rulebook_excel: Non-numeric value for proportion at row {row_num}.")
                return None
            if not (0 <= total_proportion <= 1):
                print(f"parse_rulebook_excel: Value out of [0,1] range at row {row_num} for proportion.")
                return None

            # Validate sentiment proportions
            if any(not isinstance(val, (int, float)) for val in (c, d, e)):
                print(f"parse_rulebook_excel: Non-numeric value for sentiment at row {row_num}.")
                return None
            if not all(0 <= val <= 1 for val in (c, d, e)):
                print(f"parse_rulebook_excel: Values out of [0,1] range at row {row_num} for sentiment")
                return None
            if abs((c + d + e) - 1) > 1e-9:
                print(f"parse_rulebook_excel: Sentiment does not sum to 1 at row {row_num}.")
                return None

            # Validate chunk_min_wc: must be non-None, numeric, an integer and > 0
            if chunk_min_wc is None or not isinstance(chunk_min_wc, int):
                print(f"parse_rulebook_excel: Non-integer value for chunk_min_wc at row {row_num}.")
                return None
            if chunk_min_wc <= 0:
                print(f"parse_rulebook_excel: Value for chunk_min_wc must be greater than 0 at row {row_num}.")
                return None

            # Validate chunk_max_wc: must be non-None, numeric, an integer and > chunk_min_wc
            if chunk_max_wc is None or not isinstance(chunk_max_wc, int):
                print(f"parse_rulebook_excel: Non-integer value for chunk_max_wc at row {row_num}.")
                return None
            if chunk_max_wc <= chunk_min_wc:
                print(f"parse_rulebook_excel: Value for chunk_max_wc I must be greater than chunk_min_wc at row {row_num}.")
                return None
            
            # Default chunk_pref to 0.50 if no value is present
            chunk_pref = 0.50
            # If collection_mode is "word", process chunk_pref as the preffered number of chunks
            if collection_mode == "word":
                if chunk_pref_raw is not None:
                    chunk_count_pref_str = str(chunk_pref_raw).strip()
                    mapping = {
                        "Lowest Number": get_setting('CHUNK_COUNT_MAPPING', 'Lowest_Number'),
                        "Low Number": get_setting('CHUNK_COUNT_MAPPING', 'Low_Number'),
                        "Mean Number": get_setting('CHUNK_COUNT_MAPPING', 'Mean_Number'),
                        "High Number": get_setting('CHUNK_COUNT_MAPPING', 'High_Number'),
                        "Highest Number": get_setting('CHUNK_COUNT_MAPPING', 'Highest_Number')
                    }
                    chunk_pref = mapping.get(chunk_count_pref_str, 0.5)
            
            # If collection_mode is "chunk", process chunk_pref as the preffered chunk word count size
            else:
                if chunk_pref_raw is not None:
                    chunk_count_pref_str = str(chunk_pref_raw).strip()
                    mapping = {
                        "Very Small": get_setting('CHUNK_SIZE_MAPPING', 'Very_Small'),
                        "Small": get_setting('CHUNK_SIZE_MAPPING', 'Small'),
                        "Medium": get_setting('CHUNK_SIZE_MAPPING', 'Medium'),
                        "Large": get_setting('CHUNK_SIZE_MAPPING', 'Large'),
                        "Very Large": get_setting('CHUNK_SIZE_MAPPING', 'Very_Large')
                    }
                    chunk_pref = mapping.get(chunk_count_pref_str, 0.5)
                    

            # Process chunk_wc_distribution: map expected strings to values.
            # Default to "Average Variation" if no value is present.
            chunk_wc_distribution = 5.0
            if chunk_wc_distribution_raw is not None:
                chunk_wc_distribution_str = str(chunk_wc_distribution_raw).strip()
                mapping_dist = {
                    "Low Variation": get_setting('CHUNK_VAR_MAPPING', 'Low_Variation'),
                    "Average Variation": get_setting('CHUNK_VAR_MAPPING', 'Average_Variation'),
                    "High Variation": get_setting('CHUNK_VAR_MAPPING', 'High_Variation')
                }
                chunk_wc_distribution = mapping_dist.get(chunk_wc_distribution_str, 5.0)

            # Only include rows with a positive total_proportion value.
            if total_proportion > 0:
                content_rules[topic] = {
                    'total_proportion': total_proportion,
                    'sentiment_proportion': (c, d, e),
                    'chunk_min_wc': chunk_min_wc,
                    'chunk_max_wc': chunk_max_wc,
                    'chunk_pref': chunk_pref,
                    'chunk_wc_distribution': chunk_wc_distribution
                }

        # Verify that the sum of total_proportion values equals 1
        total_prob = sum(rule['total_proportion'] for rule in content_rules.values())
        if abs(total_prob - 1) > 1e-9:
            print("parse_rulebook_excel: Sum of column B values does not equal 1.")
            return None

        # Process COLLECTION_RANGES sheet for collection_ranges
        if "COLLECTION_RANGES" not in wb.sheetnames:
            print(f"parse_rulebook_excel: 'COLLECTION_RANGES' sheet not found in workbook: {file_path}")
            return None

        COLLECTION_RANGES_ws = wb["COLLECTION_RANGES"]
        collection_ranges = []
        for row_num, row in enumerate(COLLECTION_RANGES_ws.iter_rows(min_row=4, max_col=3, values_only=True), start=4):
            # Expecting: Column A: start, Column B: end, Column C: proportion
            start_val, end_val, prop_val = row[0], row[1], row[2]
            # Terminate iteration if the start cell is empty
            if start_val is None:
                break

            # Validate that all three cells are present and numeric
            if start_val is None or end_val is None or prop_val is None:
                print(f"parse_rulebook_excel: Missing value in COLLECTION_RANGES sheet at row {row_num}.")
                return None
            if not isinstance(start_val, int) or not isinstance(end_val, int) or not isinstance(prop_val, (int, float)):
                print(f"parse_rulebook_excel: Incorrect value in COLLECTION_RANGES sheet at row {row_num}.")
                return None

            # Validate that end is greater than or equal to start
            if end_val < start_val:
                print(f"parse_rulebook_excel: In COLLECTION_RANGES sheet at row {row_num}, end ({end_val}) must be greater than start ({start_val}).")
                return None

            # If this is not the first range, ensure the current start is previous end + 1
            if collection_ranges:
                previous_end = collection_ranges[-1]['range'][1]
                if start_val != previous_end + 1:
                    print(f"parse_rulebook_excel: In COLLECTION_RANGES sheet at row {row_num}, start ({start_val}) must be equal to previous end + 1 ({previous_end + 1}).")
                    return None

            collection_ranges.append({
                'range': (int(start_val), int(end_val)),
                'target_fraction': float(prop_val)
            })

        # Validate that the sum of all proportions equals 1
        total_fraction = sum(item['target_fraction'] for item in collection_ranges)
        if abs(total_fraction - 1) > 1e-9:
            print("parse_rulebook_excel: Sum of fractions in COLLECTION_RANGES sheet does not equal 1.")
            return None

        result = {
            'collection_mode': collection_mode,
            'review_item': review_item,
            'total': total,
            'content_rules': content_rules,
            'collection_ranges': collection_ranges
        }

        # Write result to JSON in root/_data/rulebooks/json directory
        json_dir = Path(__file__).parent.parent / "_data" / "rulebooks" / "json"
        json_dir.mkdir(parents=True, exist_ok=True)
        base_filename = f"{review_item} - {collection_mode} - {total}.json"
        json_path = Path(json_dir / base_filename)
        counter = 1
        while json_path.exists():
            json_path = json_dir / f"{review_item} - {collection_mode} - {total} ({counter}).json"
            counter += 1

        try:
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(result, f, indent=4)
        except Exception as write_err:
            print(f"parse_rulebook_excel: Error writing JSON file: {write_err}")
            return None

        return json_path

    except Exception as e:
        print(f"parse_rulebook_excel: {e}")
        return None

def validate_rulebook_json(json_file_path: Path) -> Optional[Path]:
    """
    Validates a JSON file to ensure it follows the same structure and rules as produced by parse_rulebook_excel.
    
    The JSON file must have the following structure:
      - review_item: a string
      - collection_mode: either "word" or "chunk"
      - total: an integer greater than 0
      - content_rules: a dict mapping topic names to a dict with keys:
            - total_proportion: a number between 0 and 1
            - sentiment_proportion: a list or tuple of three numbers (each between 0 and 1) that sum to 1
            - chunk_min_wc: an integer > 0
            - chunk_max_wc: an integer greater than chunk_min_wc
            - chunk_pref: a number between 0 and 1
            - chunk_wc_distribution: a positive integer (validated here as a number > 0 that is integer-valued)
      - collection_ranges: a list of dicts, each with:
            - range: a list or tuple of two integers [start, end] where end >= start
              and for each subsequent range the start must equal the previous end + 1.
            - target_fraction: a number, and the sum of these fractions must equal 1.
    
    Args:
        json_file_path (Path): Path to the JSON file to validate.
        
    Returns:
        Optional[Path]: Full path to the new JSON file on success.
        None if a validation error occurs.
    """
    try:
        # Check if file exists and is a JSON file
        if not json_file_path.exists():
            print(f"validate_rulebook_json: File does not exist: {json_file_path}")
            return None
        if json_file_path.suffix.lower() != ".json":
            print(f"validate_rulebook_json: File is not a JSON file: {json_file_path}")
            return None

        # Load the JSON data
        with open(json_file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Check for required top-level keys
        required_keys = {"review_item", "collection_mode", "total", "content_rules", "collection_ranges"}
        for key in required_keys:
            if key not in data:
                print(f"validate_rulebook_json: Missing key '{key}' in JSON data.")
                return None

        # Validate review_item
        review_item = data["review_item"]
        if not isinstance(review_item, str):
            print("validate_rulebook_json: 'review_item' must be a string.")
            return None

        # Validate collection_mode
        collection_mode = data["collection_mode"]
        if collection_mode not in ("word", "chunk"):
            print("validate_rulebook_json: 'collection_mode' must be either 'word' or 'chunk'.")
            return None

        # Validate total
        total = data["total"]
        if not isinstance(total, int) or total <= 0:
            print("validate_rulebook_json: 'total' must be an integer greater than 0.")
            return None

        # Validate content_rules
        content_rules = data["content_rules"]
        if not isinstance(content_rules, dict):
            print("validate_rulebook_json: 'content_rules' must be a dictionary.")
            return None

        sum_total_proportion = 0.0
        for topic, rule in content_rules.items():
            if not isinstance(topic, str):
                print("validate_rulebook_json: All keys in 'content_rules' must be strings.")
                return None
            if not isinstance(rule, dict):
                print(f"validate_rulebook_json: The rule for topic '{topic}' must be a dictionary.")
                return None

            required_rule_keys = {
                "total_proportion",
                "sentiment_proportion",
                "chunk_min_wc",
                "chunk_max_wc",
                "chunk_pref",
                "chunk_wc_distribution"
            }
            for rkey in required_rule_keys:
                if rkey not in rule:
                    print(f"validate_rulebook_json: Missing key '{rkey}' in rule for topic '{topic}'.")
                    return None

            # total_proportion: numeric and between 0 and 1
            total_proportion = rule["total_proportion"]
            if not isinstance(total_proportion, (int, float)):
                print(f"validate_rulebook_json: 'total_proportion' for topic '{topic}' must be numeric.")
                return None
            if not (0 <= total_proportion <= 1):
                print(f"validate_rulebook_json: 'total_proportion' for topic '{topic}' must be between 0 and 1.")
                return None
            sum_total_proportion += total_proportion

            # sentiment_proportion: list/tuple of three numbers in [0,1] that sum to 1
            sentiment = rule["sentiment_proportion"]
            if not isinstance(sentiment, (list, tuple)) or len(sentiment) != 3:
                print(f"validate_rulebook_json: 'sentiment_proportion' for topic '{topic}' must be a list of three numbers.")
                return None
            for val in sentiment:
                if not isinstance(val, (int, float)):
                    print(f"validate_rulebook_json: All values in 'sentiment_proportion' for topic '{topic}' must be numeric.")
                    return None
                if not (0 <= val <= 1):
                    print(f"validate_rulebook_json: Values in 'sentiment_proportion' for topic '{topic}' must be between 0 and 1.")
                    return None
            if abs(sum(sentiment) - 1) > 1e-9:
                print(f"validate_rulebook_json: 'sentiment_proportion' values for topic '{topic}' do not sum to 1.")
                return None

            # chunk_min_wc: int > 0
            chunk_min_wc = rule["chunk_min_wc"]
            if not isinstance(chunk_min_wc, int) or chunk_min_wc <= 0:
                print(f"validate_rulebook_json: 'chunk_min_wc' for topic '{topic}' must be an integer greater than 0.")
                return None

            # chunk_max_wc: int greater than chunk_min_wc
            chunk_max_wc = rule["chunk_max_wc"]
            if not isinstance(chunk_max_wc, int) or chunk_max_wc <= chunk_min_wc:
                print(f"validate_rulebook_json: 'chunk_max_wc' for topic '{topic}' must be an integer greater than 'chunk_min_wc'.")
                return None

            # chunk_pref: numeric between 0 and 1
            chunk_pref = rule["chunk_pref"]
            if not isinstance(chunk_pref, (int, float)) or not (0 <= chunk_pref <= 1):
                print(f"validate_rulebook_json: 'chunk_pref' for topic '{topic}' must be a numeric value between 0 and 1.")
                return None

            # chunk_wc_distribution: positive number
            chunk_wc_distribution = rule["chunk_wc_distribution"]
            if not isinstance(chunk_wc_distribution, (int, float)) or chunk_wc_distribution <= 0:
                print(f"validate_rulebook_json: 'chunk_wc_distribution' for topic '{topic}' must be a positive number.")
                return None

        # Validate that the sum of total_proportion values equals 1
        if abs(sum_total_proportion - 1) > 1e-9:
            print("validate_rulebook_json: Sum of 'total_proportion' values in 'content_rules' does not equal 1.")
            return None

        # Validate collection_ranges
        collection_ranges = data["collection_ranges"]
        if not isinstance(collection_ranges, list):
            print("validate_rulebook_json: 'collection_ranges' must be a list.")
            return None

        sum_target_fraction = 0.0
        previous_end = None
        for idx, range_dict in enumerate(collection_ranges):
            if not isinstance(range_dict, dict):
                print("validate_rulebook_json: Each item in 'collection_ranges' must be a dictionary.")
                return None
            if "range" not in range_dict or "target_fraction" not in range_dict:
                print("validate_rulebook_json: Each item in 'collection_ranges' must contain 'range' and 'target_fraction'.")
                return None

            range_val = range_dict["range"]
            if not isinstance(range_val, (list, tuple)) or len(range_val) != 2:
                print("validate_rulebook_json: 'range' in collection_ranges must be a list or tuple of two integers.")
                return None
            start_val, end_val = range_val
            if not (isinstance(start_val, int) and isinstance(end_val, int)):
                print("validate_rulebook_json: 'range' values in collection_ranges must be integers.")
                return None
            if end_val < start_val:
                print("validate_rulebook_json: In collection_ranges, end value must be greater than or equal to start value.")
                return None
            if previous_end is not None and start_val != previous_end + 1:
                print(f"validate_rulebook_json: In collection_ranges at index {idx}, start value must be previous end + 1.")
                return None
            previous_end = end_val

            target_fraction = range_dict["target_fraction"]
            if not isinstance(target_fraction, (int, float)):
                print("validate_rulebook_json: 'target_fraction' in collection_ranges must be numeric.")
                return None
            sum_target_fraction += target_fraction

        if abs(sum_target_fraction - 1) > 1e-9:
            print("validate_rulebook_json: Sum of 'target_fraction' values in collection_ranges does not equal 1.")
            return None

        # Write validated JSON in root/_data/rulebooks/json directory
        json_dir = Path(__file__).parent.parent / "_data" / "rulebooks" / "json"
        json_dir.mkdir(parents=True, exist_ok=True)
        base_filename = f"{review_item} - {collection_mode} - {total}.json"
        new_json_path = json_dir / base_filename
        counter = 1
        while new_json_path.exists():
            new_json_path = json_dir / f"{review_item} - {collection_mode} - {total} ({counter}).json"
            counter += 1

        with open(new_json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)

        return new_json_path

    except Exception as e:
        print(f"validate_rulebook_json: {e}")
        return None
