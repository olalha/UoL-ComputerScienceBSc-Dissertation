import openpyxl
from pathlib import Path
from typing import Optional

from utils.settings_manager import get_setting

def parse_topic_sentiment_distribution_Excel(rulebook_name: str) -> Optional[dict]:
    """
    Parse an Excel workbook for topic sentiment distribution with additional parameters.

    This function loads an Excel workbook from the specified path and extracts:
      - 'review_item': from cell A1,
      - 'total_wc': from cell F1 (must be an int > 0),
      - 'content_rules': a dictionary mapping topic names to a dictionary containing:
          - 'total_proportion': float from col B
          - 'sentiment_proportion': a tuple of three floats from cols C, D, and E (which must sum to 1)
          - 'chunk_min_wc': int from col H (must be > 0)
          - 'chunk_max_wc': int from col I (must be > chunk_min_wc)
          - 'chunk_count_pref': mapped from col J if present
          - 'chunk_wc_distribution': mapped from col K if present
      - 'collection_wc_ranges': a list of dictionaries representing word count ranges,
          each with keys:
              - 'range': a tuple (start, end) where end > start
              - 'target_fraction': fraction for this range (must sum to 1)

    Args:
        rulebook_name (str): The file name of the Excel workbook, located in the "rulebooks" directory.

    Returns:
        Optional[dict]:
            On success, returns a dictionary with keys:
              - 'review_item'
              - 'total_wc'
              - 'content_rules'
              - 'collection_wc_ranges'
            Returns None if an error occurs.
    """
    try:
        # Get the file path relative to the script location
        file_path = Path(__file__).parent / "rulebooks" / rulebook_name

        if not file_path.exists():
            print(f"parse_topic_sentiment_distribution_Excel: File does not exist: {file_path}")
            return None
        if file_path.suffix.lower() not in {'.xlsx', '.xlsm'}:
            print(f"parse_topic_sentiment_distribution_Excel: File is not an Excel workbook: {file_path}")
            return None

        # Load the workbook and ensure the "MAIN" sheet is present
        wb = openpyxl.load_workbook(file_path)
        if "MAIN" not in wb.sheetnames:
            print(f"parse_topic_sentiment_distribution_Excel: 'MAIN' sheet not found in workbook: {file_path}")
            return None
        ws = wb["MAIN"]

        # Retrieve review_item from cell A1 and validate it
        review_item = ws['A1'].value
        if not isinstance(review_item, str):
            print("parse_topic_sentiment_distribution_Excel: Invalid value in cell A1 for review_item.")
            return None

        # Retrieve total_wc from cell F1 and validate it
        total_wc = ws['F1'].value
        if not isinstance(total_wc, int) or total_wc <= 0:
            print("parse_topic_sentiment_distribution_Excel: Invalid value in cell F1 for total_wc.")
            return None

        content_rules = {}
        # Iterate over rows starting at row 5; include columns A-E and H-K.
        for row_num, row in enumerate(ws.iter_rows(min_row=5, max_col=11, values_only=True), start=5):
            # Column A: topic (index 0)
            topic = row[0]
            # Column B: total_proportion (index 1)
            total_proportion = row[1]
            # Columns C-E: sentiment proportions (indices 2, 3, 4)
            c = row[2]
            d = row[3]
            e = row[4]
            # Column H: chunk_min_wc (index 7)
            chunk_min_wc = row[7]
            # Column I: chunk_max_wc (index 8)
            chunk_max_wc = row[8]
            # Column J: chunk_count_pref (index 9)
            chunk_count_pref_raw = row[9]
            # Column K: chunk_wc_distribution (index 10)
            chunk_wc_distribution_raw = row[10]

            # Terminate iteration if the topic cell is empty
            if topic is None:
                break
            if not isinstance(topic, str):
                print(f"parse_topic_sentiment_distribution_Excel: Non-string value in column A at row {row_num}.")
                return None

            # Validate total_proportion (col B)
            if not isinstance(total_proportion, (int, float)):
                print(f"parse_topic_sentiment_distribution_Excel: Non-numeric value in column B at row {row_num}.")
                return None
            if not (0 <= total_proportion <= 1):
                print(f"parse_topic_sentiment_distribution_Excel: Value out of [0,1] range at row {row_num} for column B.")
                return None

            # Validate sentiment proportions (cols C-E)
            if any(not isinstance(val, (int, float)) for val in (c, d, e)):
                print(f"parse_topic_sentiment_distribution_Excel: Non-numeric value in columns C-E at row {row_num}.")
                return None
            if not all(0 <= val <= 1 for val in (c, d, e)):
                print(f"parse_topic_sentiment_distribution_Excel: Values out of [0,1] range at row {row_num} for columns C-E.")
                return None
            if abs((c + d + e) - 1) > 1e-9:
                print(f"parse_topic_sentiment_distribution_Excel: Columns C-E do not sum to 1 at row {row_num}.")
                return None

            # Validate chunk_min_wc (col H): must be non-None, numeric, an integer and > 0
            if chunk_min_wc is None or not isinstance(chunk_min_wc, int):
                print(f"parse_topic_sentiment_distribution_Excel: Non-integer value for chunk_min_wc in column H at row {row_num}.")
                return None
            if chunk_min_wc <= 0:
                print(f"parse_topic_sentiment_distribution_Excel: Value in column H must be greater than 0 at row {row_num}.")
                return None

            # Validate chunk_max_wc (col I): must be non-None, numeric, an integer and > chunk_min_wc
            if chunk_max_wc is None or not isinstance(chunk_max_wc, int):
                print(f"parse_topic_sentiment_distribution_Excel: Non-integer value for chunk_max_wc in column I at row {row_num}.")
                return None
            if chunk_max_wc <= chunk_min_wc:
                print(f"parse_topic_sentiment_distribution_Excel: Value in column I must be greater than chunk_min_wc at row {row_num}.")
                return None

            # Process chunk_count_pref (col J): map expected strings to values.
            # Default to "Mean Number" if no value is present.
            chunk_count_pref = 0.50
            if chunk_count_pref_raw is not None:
                chunk_count_pref_str = str(chunk_count_pref_raw).strip()
                mapping = {
                    "Lowest Number": get_setting('CHUNK_COUNT_MAPPING', 'Lowest_Number'),
                    "Low Number": get_setting('CHUNK_COUNT_MAPPING', 'Low_Number'),
                    "Mean Number": get_setting('CHUNK_COUNT_MAPPING', 'Mean_Number'),
                    "High Number": get_setting('CHUNK_COUNT_MAPPING', 'High_Number'),
                    "Highest Number": get_setting('CHUNK_COUNT_MAPPING', 'Highest_Number')
                }
                chunk_count_pref = mapping.get(chunk_count_pref_str, 0.5)

            # Process chunk_wc_distribution (col K): map expected strings to values.
            # Default to "Average Variation" if no value is present.
            chunk_wc_distribution = 5.0
            if chunk_wc_distribution_raw is not None:
                chunk_wc_distribution_str = str(chunk_wc_distribution_raw).strip()
                mapping_dist = {
                    "Low Variation": get_setting('CHUNK_VAR_MAPPING', 'Low_Variation'),
                    "Average Variation": get_setting('CHUNK_VAR_MAPPING', 'Average_Variation'),
                    "High Variation": get_setting('CHUNK_VAR_MAPPING', 'High_Variation')
                }
                chunk_wc_distribution = mapping_dist.get(chunk_wc_distribution_str, 5.0)

            # Only include rows with a positive total_proportion value.
            if total_proportion > 0:
                content_rules[topic] = {
                    'total_proportion': total_proportion,
                    'sentiment_proportion': (c, d, e),
                    'chunk_min_wc': chunk_min_wc,
                    'chunk_max_wc': chunk_max_wc,
                    'chunk_count_pref': chunk_count_pref,
                    'chunk_wc_distribution': chunk_wc_distribution
                }

        # Verify that the sum of total_proportion values equals 1
        total_prob = sum(rule['total_proportion'] for rule in content_rules.values())
        if abs(total_prob - 1) > 1e-9:
            print("parse_topic_sentiment_distribution_Excel: Sum of column B values does not equal 1.")
            return None

        # Process ALLOCATION sheet for collection_wc_ranges
        if "ALLOCATION" not in wb.sheetnames:
            print(f"parse_topic_sentiment_distribution_Excel: 'ALLOCATION' sheet not found in workbook: {file_path}")
            return None

        allocation_ws = wb["ALLOCATION"]
        collection_wc_ranges = []
        for row_num, row in enumerate(allocation_ws.iter_rows(min_row=4, max_col=3, values_only=True), start=4):
            # Expecting: Column A: start, Column B: end, Column C: proportion
            start_val, end_val, prop_val = row[0], row[1], row[2]
            # Terminate iteration if the start cell is empty
            if start_val is None:
                break

            # Validate that all three cells are present and numeric
            if start_val is None or end_val is None or prop_val is None:
                print(f"parse_topic_sentiment_distribution_Excel: Missing value in ALLOCATION sheet at row {row_num}.")
                return None
            if not isinstance(start_val, int) or not isinstance(end_val, int) or not isinstance(prop_val, (int, float)):
                print(f"parse_topic_sentiment_distribution_Excel: Incorrect value in ALLOCATION sheet at row {row_num}.")
                return None

            # Validate that end is greater than start
            if end_val <= start_val:
                print(f"parse_topic_sentiment_distribution_Excel: In ALLOCATION sheet at row {row_num}, end ({end_val}) must be greater than start ({start_val}).")
                return None

            # If this is not the first range, ensure the current start is previous end + 1
            if collection_wc_ranges:
                previous_end = collection_wc_ranges[-1]['range'][1]
                if start_val != previous_end + 1:
                    print(f"parse_topic_sentiment_distribution_Excel: In ALLOCATION sheet at row {row_num}, start ({start_val}) must be equal to previous end + 1 ({previous_end + 1}).")
                    return None

            collection_wc_ranges.append({
                'range': (int(start_val), int(end_val)),
                'target_fraction': float(prop_val)
            })

        # Validate that the sum of all proportions equals 1
        total_fraction = sum(item['target_fraction'] for item in collection_wc_ranges)
        if abs(total_fraction - 1) > 1e-9:
            print("parse_topic_sentiment_distribution_Excel: Sum of fractions in ALLOCATION sheet does not equal 1.")
            return None

        return {
            'review_item': review_item,
            'total_wc': total_wc,
            'content_rules': content_rules,
            'collection_wc_ranges': collection_wc_ranges
        }

    except Exception as e:
        print(f"parse_topic_sentiment_distribution_Excel: {e}")
        return None
