import openpyxl
import json
from pathlib import Path
from typing import Optional

from utils.settings_manager import get_setting

def parse_rulebook_excel(file_path: Path) -> Optional[dict]:
    """
    Parses an Excel workbook defining the topic sentiment distribution for a text corpus with additional parameters.
    
    The Excel workbook must adhere to the template structure, including:
      - A "MAIN" sheet containing:
          - 'content_title' in cell A1 (string)
          - 'collection_mode' in cell E1 (either "word" or "chunk")
          - 'total' in cell G1 (an integer greater than 0)
          - 'content_rules' starting from row 5 with columns:
                * Column A: topic (string)
                * Column B: total_proportion (numeric value between 0 and 1)
                * Columns C-E: sentiment_proportion (three numbers that sum to 1)
                * Column H: chunk_min_wc (integer > 0)
                * Column I: chunk_max_wc (integer greater than chunk_min_wc)
                * Column J: chunk_pref (preferred number based on collection_mode)
                * Column K: chunk_wc_distribution (optional, with a default if not provided)
      - A "COLLECTION_RANGES" sheet containing:
          - Ranges defined with start, end, and target_fraction (with contiguous ranges and sum of target_fraction equal to 1)
    
    After extracting and mapping the values, the function validates the data using a shared helper.
    If validation passes, the rulebook is written to a JSON file in the "_data/rulebooks/json" directory.

    Args:
        file_path (Path): Full path to the Excel workbook.

    Returns:
        Optional[Path]: Full path to the newly created JSON file on success, or None if a validation or processing error occurs.
    """
    try:
        # Validate the file path and extension
        if not file_path.exists():
            print(f"parse_rulebook_excel: File does not exist: {file_path}")
            return None
        if file_path.suffix.lower() not in {'.xlsx', '.xlsm'}:
            print(f"parse_rulebook_excel: File is not an Excel workbook: {file_path}")
            return None

        wb = openpyxl.load_workbook(file_path)
        if "MAIN" not in wb.sheetnames:
            print(f"parse_rulebook_excel: 'MAIN' sheet not found in workbook: {file_path}")
            return None
        ws = wb["MAIN"]

        # Retrieve top-level values
        content_title = ws['A1'].value
        collection_mode = ws['E1'].value
        total = ws['G1'].value

        # Validate top-level values
        if not isinstance(content_title, str):
            print("parse_rulebook_excel: Invalid value in cell A1 for content_title.")
            return None
        if collection_mode not in ('word', 'chunk'):
            print("parse_rulebook_excel: Invalid value in cell E1 for collection_mode.")
            return None
        if not isinstance(total, int) or total <= 0:
            print("parse_rulebook_excel: Invalid value in cell G1 for total.")
            return None

        content_rules = {}
        # Process rows from the MAIN sheet starting at row 5
        for row_num, row in enumerate(ws.iter_rows(min_row=5, max_col=11, values_only=True), start=5):
            topic = row[0]
            if topic is None:
                break
            if not isinstance(topic, str):
                print(f"parse_rulebook_excel: Non-string value for topic at row {row_num}.")
                return None

            # Extract values from the row
            total_proportion = row[1]
            c, d, e = row[2], row[3], row[4]
            chunk_min_wc = row[7]
            chunk_max_wc = row[8]
            chunk_pref_raw = row[9]
            chunk_wc_distribution_raw = row[10]

            # Validate and process the values
            if not isinstance(total_proportion, (int, float)):
                print(f"parse_rulebook_excel: Non-numeric value for proportion at row {row_num}.")
                return None
            if any(not isinstance(val, (int, float)) for val in (c, d, e)):
                print(f"parse_rulebook_excel: Non-numeric value for sentiment at row {row_num}.")
                return None
            if chunk_min_wc is None or not isinstance(chunk_min_wc, int):
                print(f"parse_rulebook_excel: Non-integer value for chunk_min_wc at row {row_num}.")
                return None
            if chunk_max_wc is None or not isinstance(chunk_max_wc, int):
                print(f"parse_rulebook_excel: Non-integer value for chunk_max_wc at row {row_num}.")
                return None

            # Process chunk_pref based on collection_mode
            chunk_pref = 0.50
            if collection_mode == "word":
                if chunk_pref_raw is not None:
                    chunk_count_pref_str = str(chunk_pref_raw).strip()
                    mapping = {
                        "Lowest Number": get_setting('CHUNK_COUNT_MAPPING', 'Lowest_Number'),
                        "Low Number": get_setting('CHUNK_COUNT_MAPPING', 'Low_Number'),
                        "Mean Number": get_setting('CHUNK_COUNT_MAPPING', 'Mean_Number'),
                        "High Number": get_setting('CHUNK_COUNT_MAPPING', 'High_Number'),
                        "Highest Number": get_setting('CHUNK_COUNT_MAPPING', 'Highest_Number')
                    }
                    chunk_pref = mapping.get(chunk_count_pref_str, 0.5)
            else:
                if chunk_pref_raw is not None:
                    chunk_count_pref_str = str(chunk_pref_raw).strip()
                    mapping = {
                        "Very Small": get_setting('CHUNK_SIZE_MAPPING', 'Very_Small'),
                        "Small": get_setting('CHUNK_SIZE_MAPPING', 'Small'),
                        "Medium": get_setting('CHUNK_SIZE_MAPPING', 'Medium'),
                        "Large": get_setting('CHUNK_SIZE_MAPPING', 'Large'),
                        "Very Large": get_setting('CHUNK_SIZE_MAPPING', 'Very_Large')
                    }
                    chunk_pref = mapping.get(chunk_count_pref_str, 0.5)

            # Process chunk_wc_distribution with default value if not provided
            chunk_wc_distribution = 5.0
            if chunk_wc_distribution_raw is not None:
                chunk_wc_distribution_str = str(chunk_wc_distribution_raw).strip()
                mapping_dist = {
                    "Low Variation": get_setting('CHUNK_VAR_MAPPING', 'Low_Variation'),
                    "Average Variation": get_setting('CHUNK_VAR_MAPPING', 'Average_Variation'),
                    "High Variation": get_setting('CHUNK_VAR_MAPPING', 'High_Variation')
                }
                chunk_wc_distribution = mapping_dist.get(chunk_wc_distribution_str, 5.0)

            if total_proportion > 0:
                content_rules[topic] = {
                    'total_proportion': total_proportion,
                    'sentiment_proportion': (c, d, e),
                    'chunk_min_wc': chunk_min_wc,
                    'chunk_max_wc': chunk_max_wc,
                    'chunk_pref': chunk_pref,
                    'chunk_wc_distribution': chunk_wc_distribution
                }

        # Process COLLECTION_RANGES sheet
        if "COLLECTION_RANGES" not in wb.sheetnames:
            print(f"parse_rulebook_excel: 'COLLECTION_RANGES' sheet not found in workbook: {file_path}")
            return None

        # Validate and process collection ranges
        COLLECTION_RANGES_ws = wb["COLLECTION_RANGES"]
        collection_ranges = []
        for row_num, row in enumerate(COLLECTION_RANGES_ws.iter_rows(min_row=4, max_col=3, values_only=True), start=4):
            start_val, end_val, prop_val = row[0], row[1], row[2]
            if start_val is None:
                break
            if start_val is None or end_val is None or prop_val is None:
                print(f"parse_rulebook_excel: Missing value in COLLECTION_RANGES sheet at row {row_num}.")
                return None
            if not isinstance(start_val, int) or not isinstance(end_val, int):
                print(f"parse_rulebook_excel: Incorrect value type in COLLECTION_RANGES sheet at row {row_num}.")
                return None

            collection_ranges.append({
                'range': (int(start_val), int(end_val)),
                'target_fraction': float(prop_val)
            })

        result = {
            'collection_mode': collection_mode,
            'content_title': content_title,
            'total': total,
            'content_rules': content_rules,
            'collection_ranges': collection_ranges
        }

        # Validate the merged rulebook values
        if not validate_rulebook_values(result):
            return None

        # Return the validated rulebook if successful
        return result

    except Exception as e:
        print(f"parse_rulebook_excel: {e}")
        return None

def validate_rulebook_values(rulebook: dict) -> bool:
    """
    Validates a rulebook dictionary to ensure it conforms to the expected structure and value constraints.

    The rulebook dictionary must include the following keys:
      - content_title: a string
      - collection_mode: either "word" or "chunk"
      - total: an integer greater than 0
      - content_rules: a dictionary mapping topic names to a dictionary with keys:
            - total_proportion: a number between 0 and 1
            - sentiment_proportion: a list or tuple of three numbers (each between 0 and 1) that sum to 1
            - chunk_min_wc: an integer greater than 0
            - chunk_max_wc: an integer greater than chunk_min_wc
            - chunk_pref: a number between 0 and 1
            - chunk_wc_distribution: a positive number (validated as a number > 0)
      - collection_ranges: a list of dictionaries, each containing:
            - range: a list or tuple of two integers [start, end] where end >= start and for each subsequent range,
              the start value must equal the previous end + 1.
            - target_fraction: a number, and the sum of these target fractions must equal 1.

    Args:
        rulebook (dict): A dictionary representing the rulebook extracted from Excel or JSON.

    Returns:
        bool: True if the rulebook values are valid, False otherwise.
    """
    # Validate the rulebook dictionary
    if not rulebook or not isinstance(rulebook, dict):
        print(f"validate_rulebook_values: Rulebook data is empty or not a dictionary.")
        return False
    
    # Validate top-level keys
    required_keys = {"content_title", "collection_mode", "total", "content_rules", "collection_ranges"}
    for key in required_keys:
        if key not in rulebook:
            print(f"validate_rulebook_values: Missing key '{key}' in rulebook data.")
            return False

    # Validate content_title
    if not isinstance(rulebook["content_title"], str):
        print("validate_rulebook_values: 'content_title' must be a string.")
        return False

    # Validate collection_mode
    if rulebook["collection_mode"] not in ("word", "chunk"):
        print("validate_rulebook_values: 'collection_mode' must be either 'word' or 'chunk'.")
        return False

    # Validate total
    if not isinstance(rulebook["total"], int) or rulebook["total"] <= 0:
        print("validate_rulebook_values: 'total' must be an integer greater than 0.")
        return False

    # Validate content_rules
    content_rules = rulebook["content_rules"]
    if not isinstance(content_rules, dict):
        print("validate_rulebook_values: 'content_rules' must be a dictionary.")
        return False

    sum_total_proportion = 0.0
    for topic, rule in content_rules.items():
        if not isinstance(topic, str):
            print("validate_rulebook_values: All keys in 'content_rules' must be strings.")
            return False
        if not isinstance(rule, dict):
            print(f"validate_rulebook_values: The rule for topic '{topic}' must be a dictionary.")
            return False

        required_rule_keys = {
            "total_proportion",
            "sentiment_proportion",
            "chunk_min_wc",
            "chunk_max_wc",
            "chunk_pref",
            "chunk_wc_distribution"
        }
        for rkey in required_rule_keys:
            if rkey not in rule:
                print(f"validate_rulebook_values: Missing key '{rkey}' in rule for topic '{topic}'.")
                return False

        # total_proportion: numeric and between 0 and 1
        total_proportion = rule["total_proportion"]
        if not isinstance(total_proportion, (int, float)):
            print(f"validate_rulebook_values: 'total_proportion' for topic '{topic}' must be numeric.")
            return False
        if not (0 <= total_proportion <= 1):
            print(f"validate_rulebook_values: 'total_proportion' for topic '{topic}' must be between 0 and 1.")
            return False
        sum_total_proportion += total_proportion

        # sentiment_proportion: list/tuple of three numbers in [0,1] that sum to 1
        sentiment = rule["sentiment_proportion"]
        if not isinstance(sentiment, (list, tuple)) or len(sentiment) != 3:
            print(f"validate_rulebook_values: 'sentiment_proportion' for topic '{topic}' must be a list or tuple of three numbers.")
            return False
        for val in sentiment:
            if not isinstance(val, (int, float)):
                print(f"validate_rulebook_values: All values in 'sentiment_proportion' for topic '{topic}' must be numeric.")
                return False
            if not (0 <= val <= 1):
                print(f"validate_rulebook_values: Values in 'sentiment_proportion' for topic '{topic}' must be between 0 and 1.")
                return False
        if abs(sum(sentiment) - 1) > 1e-9:
            print(f"validate_rulebook_values: 'sentiment_proportion' values for topic '{topic}' do not sum to 1.")
            return False

        # chunk_min_wc: int > 0
        chunk_min_wc = rule["chunk_min_wc"]
        if not isinstance(chunk_min_wc, int) or chunk_min_wc <= 0:
            print(f"validate_rulebook_values: 'chunk_min_wc' for topic '{topic}' must be an integer greater than 0.")
            return False

        # chunk_max_wc: int > chunk_min_wc
        chunk_max_wc = rule["chunk_max_wc"]
        if not isinstance(chunk_max_wc, int) or chunk_max_wc <= chunk_min_wc:
            print(f"validate_rulebook_values: 'chunk_max_wc' for topic '{topic}' must be an integer greater than 'chunk_min_wc'.")
            return False

        # chunk_pref: numeric between 0 and 1
        chunk_pref = rule["chunk_pref"]
        if not isinstance(chunk_pref, (int, float)) or not (0 <= chunk_pref <= 1):
            print(f"validate_rulebook_values: 'chunk_pref' for topic '{topic}' must be a numeric value between 0 and 1.")
            return False

        # chunk_wc_distribution: positive number
        chunk_wc_distribution = rule["chunk_wc_distribution"]
        if not isinstance(chunk_wc_distribution, (int, float)) or chunk_wc_distribution <= 0:
            print(f"validate_rulebook_values: 'chunk_wc_distribution' for topic '{topic}' must be a positive number.")
            return False

    if abs(sum_total_proportion - 1) > 1e-9:
        print("validate_rulebook_values: Sum of 'total_proportion' values in 'content_rules' does not equal 1.")
        return False

    # Validate collection_ranges
    collection_ranges = rulebook["collection_ranges"]
    if not isinstance(collection_ranges, list):
        print("validate_rulebook_values: 'collection_ranges' must be a list.")
        return False

    sum_target_fraction = 0.0
    previous_end = None
    for idx, range_dict in enumerate(collection_ranges):
        if not isinstance(range_dict, dict):
            print("validate_rulebook_values: Each item in 'collection_ranges' must be a dictionary.")
            return False
        if "range" not in range_dict or "target_fraction" not in range_dict:
            print("validate_rulebook_values: Each item in 'collection_ranges' must contain 'range' and 'target_fraction'.")
            return False

        range_val = range_dict["range"]
        if not isinstance(range_val, (list, tuple)) or len(range_val) != 2:
            print("validate_rulebook_values: 'range' in collection_ranges must be a list or tuple of two integers.")
            return False
        start_val, end_val = range_val
        if not (isinstance(start_val, int) and isinstance(end_val, int)):
            print("validate_rulebook_values: 'range' values in collection_ranges must be integers.")
            return False
        if end_val < start_val:
            print("validate_rulebook_values: In collection_ranges, end value must be greater than or equal to start value.")
            return False
        if previous_end is not None and start_val != previous_end + 1:
            print(f"validate_rulebook_values: In collection_ranges at index {idx}, start value must be previous end + 1.")
            return False
        previous_end = end_val

        target_fraction = range_dict["target_fraction"]
        if not isinstance(target_fraction, (int, float)):
            print("validate_rulebook_values: 'target_fraction' in collection_ranges must be numeric.")
            return False
        sum_target_fraction += target_fraction

    if abs(sum_target_fraction - 1) > 1e-9:
        print("validate_rulebook_values: Sum of 'target_fraction' values in collection_ranges does not equal 1.")
        return False

    return True
