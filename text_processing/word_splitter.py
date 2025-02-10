import math
import random
import openpyxl
from pathlib import Path
import matplotlib.pyplot as plt

# TODO: Add these to settings file
CHUNK_WC_MIN = 20
CHUNK_WC_MAX = 150
TOTAL_WC = 20000

def _random_partition(N, min_val, max_val, i):
    """
    Partitions a postive int N into i random chunks within specified minimum and maximum values.
    Args:
        N (int): The total value to be partitioned
        min_val (int): Minimum allowed chunk size
        max_val (int): Maximum allowed chunk size
        i (int): The number of chunks to create
    Returns:
        list: A list of i positive integers that sum to N, each between min_val and max_val
    Note:
        - The order of the chunks is randomized
    """

    # Start with minimum values
    chunks = [min_val] * i
    # Calculate remainder to distribute
    R = N - i * min_val
    # Max additional per chunk
    max_extra = max_val - min_val

    # Distribute remainder
    for j in range(i):
        remaining_chunks = i - j - 1
        
        # Calculate min allowed for this chunk (consider remaining chunks)
        allowed_min = max(0, R - remaining_chunks * max_extra)
        # Calculate max allowed for this chunk
        allowed_max = min(max_extra, R)
        
        # Add random amount to chunk and subtract from remainder
        extra = random.randint(allowed_min, allowed_max)
        chunks[j] += extra
        R -= extra

    # Randomize order
    random.shuffle(chunks)
    return chunks

def get_chunks(N, min_val, max_val):
    """
    Find optimal number of chunks to split a postive int N into given min and max chunk size constraints.
    Given a total number N, randomly partition N into i chunks such that each chunk is between min_val and max_val.
    Args:
        N (int): Total number to split into chunks
        min_val (int): Minimum allowed chunk size
        max_val (int): Maximum allowed chunk size
    Returns:
        list: A list of i positive integers that sum to N, each between min_val and max_val.
    """
    
    # Check if the inputs are valid
    if not isinstance(N, int) or N <= 0:
        return None
    if min_val > max_val:
        return None

    # Compute the feasible range for the number of chunks.
    i_lower = math.ceil(N / max_val)
    i_upper = N // min_val

    # Check if a legal partition exists.
    if i_lower > i_upper:
        return None

    # Find the number of chunks that minimizes the
    # different between avg chunk size and target avg
    target_avg = (min_val + max_val) / 2.0
    best_i = None
    best_diff = float('inf')
    for i in range(i_lower, i_upper + 1):
        diff = abs(N / i - target_avg)
        if diff < best_diff:
            best_diff = diff
            best_i = i

    return _random_partition(N, min_val, max_val, best_i)

def allocate_chunk_collections_random(
                           chunks,
                           max_chunks_per_collection=1,
                           max_single_chunk_collection_wc=None,
                           prioritize_min_num_of_collections=0.5,
                           invalid_standalone_sentiments_collections=[]):
    
    # Check if max_chunks_per_collection is valid
    if max_chunks_per_collection < 1:
        print("allocate_chunks_random: max_chunks must be at least 1.")
        return None
    # Check if max_single_chunk_collection_wc is valid
    if max_single_chunk_collection_wc is None:
        max_single_chunk_collection_wc = CHUNK_WC_MAX
    else:
        if max_single_chunk_collection_wc < 0:
            print("allocate_chunks_random: max_single_chunk_collection_wc must be non-negative.")
            return None
        if max_single_chunk_collection_wc > CHUNK_WC_MAX:
            print("allocate_chunks_random: max_single_chunk_collection_wc exceeds CHUNK_WC_MAX.")
            return None
    # Check if prioritize_min_num_of_collections is valid
    if prioritize_min_num_of_collections < 0 or prioritize_min_num_of_collections > 1:
        print("allocate_chunks_random: prioritize_min_num_of_collections must be in [0,1].")
        return None
    # Check if invalid_standalone_sentiments_collections is valid
    sentiments_values = ["pos","neu","neg"]
    if any(s not in sentiments_values for s in invalid_standalone_sentiments_collections):
        print("allocate_chunks_random: Invalid definition of invalid_standalone_sentiments_collections list.")
        return None
    if len(invalid_standalone_sentiments_collections) > 3:
        print("allocate_chunks_random: invalid_standalone_sentiments_collections contains too many sentiments.")
        return None
    
    # Helper function to check if a chunk can be added to a collection
    def is_valid_collection(collection, chunk):
        if collection == []:
            # Check if chunk is too large to stand alone
            if chunk[2] > max_single_chunk_collection_wc:
                return False
            # Check if chunk has invalid sentiment to stand alone
            if chunk[1] in invalid_standalone_sentiments_collections:
                return False
            return True
        else:
            proposed_collection = collection + [chunk]
            # Check if collection already contains topic
            if chunk[0] in [c[0] for c in collection]:
                return False
            # Check if collection has too many chunks
            if len(proposed_collection) > max_chunks_per_collection:
                return False
            return True

    # Number of chunks
    N = len(chunks)
    
    print(N)

    # Keep track of each chunk's partner index, or -1 if none
    partners = [-1] * N

    invalid_standalone_chunks = [
        i for i in range(N)
        if not is_valid_collection([], chunks[i])
    ]

    def backtrack_invalid(i):
        """
        Attempts to pair invalid chunks from index i onward.
        Returns True if a valid overall pairing is found, otherwise False.
        """
        # If we've processed all invalid chunks, we're done
        if i == len(invalid_standalone_chunks):
            return True

        seeker_idx = invalid_standalone_chunks[i]

        # If this chunk was already paired by previous steps, move on
        if partners[seeker_idx] != -1:
            return backtrack_invalid(i + 1)

        # Try pairing this invalid chunk with any other unpaired chunk
        for candidate_idx in range(N):
            if candidate_idx != seeker_idx and partners[candidate_idx] == -1:
                # Check if the pair is valid
                if is_valid_collection([chunks[seeker_idx]], chunks[candidate_idx]):
                    # Mark them as paired
                    partners[seeker_idx] = candidate_idx
                    partners[candidate_idx] = seeker_idx

                    # Recurse to see if the rest can be paired
                    if backtrack_invalid(i + 1):
                        return True

                    # Backtrack if pairing didn't work out
                    partners[seeker_idx] = -1
                    partners[candidate_idx] = -1

        # If no valid pair could be found for this chunk, fail
        return False

    # Now, run the backtracking
    if not backtrack_invalid(0):
        print("Unable to pair all invalid standalone chunks.")
        return None
    
    all_collections = []
    curretly_unallocated = []
    
    visited_pairs = set()
    for i in range(len(partners)):
        if i not in visited_pairs:
            if partners[i] >= 0:
                # Pair found, add it and mark both as visited
                all_collections.append([chunks[i], chunks[partners[i]]])
                visited_pairs.add(partners[i])
            else:
                # Standalone chunk (valid on its own)
                curretly_unallocated.append(chunks[i])
    
    # Allocate remaining chunks to collections
    for chunk in curretly_unallocated:
        # Check if chunk can be added to existing collection
        create_new_collection = True
        if random.random() > prioritize_min_num_of_collections:
            # Add to existing collection
            for collection in all_collections:
                if is_valid_collection(collection, chunk):
                    collection.append(chunk)
                    random.shuffle(all_collections)
                    create_new_collection = False
                    break
        if create_new_collection:
                all_collections.append([chunk])
        
    return all_collections

def parse_topic_sentiment_distribution_Excel(path: str) -> dict:
    
    file_path = Path(path)
    if not file_path.exists():
        print(f"File does not exist: {path}")
        return {}
    if not file_path.suffix.lower() in ('.xlsx', '.xlsm'):
        print(f"File is not an Excel workbook: {path}")
        return {}
    wb = openpyxl.load_workbook(path)
    
    if "MAIN" not in wb.sheetnames:
        print(f"Error opening workbook or sheet: {e}")
        return {}
    ws = wb["MAIN"]

    result = {}
    row = 4
    while True:
        try:
            cell_a = ws.cell(row=row, column=1).value
        except Exception as e:
            print(f"Error accessing cell A{row}: {e}")
            return {}
        if cell_a is None:
            break
        if not isinstance(cell_a, str):
            print(f"Non-string value in column A at row {row}.")
            return {}

        cell_b = ws.cell(row=row, column=2).value
        cell_c = ws.cell(row=row, column=3).value
        cell_d = ws.cell(row=row, column=4).value
        cell_e = ws.cell(row=row, column=5).value

        if any(not isinstance(x, (int, float)) for x in [cell_b, cell_c, cell_d, cell_e]):
            print(f"Non-numeric value in columns B-E at row {row}.")
            return {}
        if not (0 <= cell_b <= 1):
            print(f"Value out of [0,1] range at row {row} for column B.")
            return {}
        if not (0 <= cell_c <= 1 and 0 <= cell_d <= 1 and 0 <= cell_e <= 1):
            print(f"Values out of [0,1] range at row {row} for columns C-E.")
            return {}
        if abs(cell_c + cell_d + cell_e - 1) > 1e-9:
            print(f"Columns C-E do not sum to 1 at row {row}.")
            return {}

        if cell_b > 0:
            result[cell_a] = (cell_b, (cell_c, cell_d, cell_e))
        
        row += 1

    total_b = sum(item[0] for item in result.values())
    if abs(total_b - 1) > 1e-9:
        print("Sum of column B values does not equal 1.")
        return {}

    return result

import matplotlib.pyplot as plt

def visualize_chunk_allocation(all_chunks):
    # Define sentiment colors
    sentiment_colors = {
        'neg': '#ff9999',  # pastel red
        'neu': '#ffff99',  # yellow
        'pos': '#99ccff'   # pastel blue
    }

    # Sort the reviews by total word count (ascending)
    all_chunks = sorted(all_chunks, key=lambda review: sum(word_count for (_, _, word_count) in review))
    
    # Define a custom order for sentiments and sort each review accordingly.
    sentiment_order = {'neg': 0, 'neu': 1, 'pos': 2}
    for review in all_chunks:
        review.sort(key=lambda tup: sentiment_order.get(tup[1], 99))
    
    # Create the plot with flipped axes: reviews on the x-axis, word count on the y-axis.
    fig, ax = plt.subplots(figsize=(10, 6))
    bar_width = 0.8  # default width for each bar

    # Loop over reviews to plot each as a vertical stacked bar.
    for i, review in enumerate(all_chunks):
        bottom = 0  # starting y-coordinate for the stacked segments in this review
        for (topic, sentiment, word_count) in review:
            # Draw one segment (vertical bar) for this chunk.
            # We set edgecolor='none' so we can manually add only top and bottom borders.
            ax.bar(i, word_count, bottom=bottom, width=bar_width,
                   color=sentiment_colors.get(sentiment, 'gray'),
                   edgecolor='none')
            
            # Compute the left and right edges of the bar for drawing horizontal lines.
            left_edge = i - bar_width/2
            right_edge = i + bar_width/2
            
            # Draw a 1px horizontal line at the bottom of the segment.
            ax.hlines(y=bottom, xmin=left_edge, xmax=right_edge, colors='black', linewidth=1)
            # Draw a 1px horizontal line at the top of the segment.
            ax.hlines(y=bottom + word_count, xmin=left_edge, xmax=right_edge, colors='black', linewidth=1)
            
            # Update bottom for the next segment.
            bottom += word_count

    # Set x-ticks and labels: now just i (starting from 0)
    ax.set_xlabel('Review')
    ax.set_ylabel('Word Count')
    ax.set_title('Stacked Bar Chart of Reviews by Sentiment (Sorted by Total Word Count)')

    plt.tight_layout()
    plt.show()

if __name__ == "__main__":

    # Get the path relative to the script location
    file_path = Path(__file__).parent / "rulebooks" / "TEMPLATE.xlsx"
    
    ts_distribution_percentage = parse_topic_sentiment_distribution_Excel(file_path)
    if not ts_distribution_percentage:
        print("Error parsing topic-sentiment distribution Excel file.")
    else:
        
        all_chunks = []
        for topic_name, topic_values in ts_distribution_percentage.items():
            topic_wc = int(TOTAL_WC * topic_values[0])
            
            for index, sentiment in enumerate(["pos", "neu", "neg"]):
                topic_sentiment_wc = int(topic_wc * topic_values[1][index])
                
                if topic_sentiment_wc == 0:
                    continue
                chunks = get_chunks(topic_sentiment_wc, CHUNK_WC_MIN, CHUNK_WC_MAX)
            
                if not chunks:
                    print(f"Error partitioning - topic:'{topic_name}' sentiment:'{sentiment}' - wc:{topic_sentiment_wc}.")
                    continue
                else:
                    for i in chunks:
                        all_chunks.append((topic_name, sentiment, i))
        
        all_reviews = allocate_chunk_collections_random(chunks=all_chunks, 
                                                        max_chunks_per_collection=10, 
                                                        max_single_chunk_collection_wc=40, 
                                                        prioritize_min_num_of_collections=1,
                                                        invalid_standalone_sentiments_collections=["neu"])
        
        if not all_reviews:
            print("Error allocating chunks to collections.")
        else:
            visualize_chunk_allocation(all_reviews)
        