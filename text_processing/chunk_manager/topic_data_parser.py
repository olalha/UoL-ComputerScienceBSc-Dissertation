import openpyxl
from pathlib import Path
from typing import Optional

def parse_topic_sentiment_distribution_Excel(rulebook_name: str) -> Optional[dict]:
    """
    Parse an Excel workbook for topic sentiment distribution.

    This function loads an Excel workbook from the specified path and extracts topic sentiment distribution
    data from a sheet named "MAIN". It processes rows beginning at row 4, expecting column A to contain topic
    names (strings) and columns B through E to contain numeric values in the range [0, 1]. For each row, it 
    verifies that columns C, D, and E sum to 1, and finally checks that the total of column B values across all 
    included rows equals 1.

    Args:
        path (str): The file path to an Excel workbook. The file must exist and have a ".xlsx" or ".xlsm" extension.

    Returns:
        Optional[dict]: A dictionary mapping topic names (str) to a tuple containing a probability (float from column B)
                        and a tuple of three floats (from columns C, D, and E). Returns None if an error occurs.
    """
    try:
         # Get the path relative to the script location
        file_path = Path(__file__).parent.parent / "rulebooks" / rulebook_name
        
        if not file_path.exists():
            print(f"parse_topic_sentiment_distribution_Excel: File does not exist: {file_path}")
            return None
        if file_path.suffix.lower() not in {'.xlsx', '.xlsm'}:
            print(f"parse_topic_sentiment_distribution_Excel: File is not an Excel workbook: {file_path}")
            return None

        # Load the workbook and ensure the "MAIN" sheet is present
        wb = openpyxl.load_workbook(file_path)
        if "MAIN" not in wb.sheetnames:
            print(f"parse_topic_sentiment_distribution_Excel: 'MAIN' sheet not found in workbook: {file_path}")
            return None
        ws = wb["MAIN"]

        result = {}
        # Iterate over rows starting at row 4; stop at the first row where column A is empty
        for row_num, row in enumerate(ws.iter_rows(min_row=4, max_col=5, values_only=True), start=4):
            # Unpack columns A through E
            topic, prob, c, d, e = row
            
            # Terminate if the topic cell is empty
            if topic is None:
                break
            if not isinstance(topic, str):
                print(f"parse_topic_sentiment_distribution_Excel: Non-string value in column A at row {row_num}.")
                return None
            if any(not isinstance(val, (int, float)) for val in (prob, c, d, e)):
                print(f"parse_topic_sentiment_distribution_Excel: Non-numeric value in columns B-E at row {row_num}.")
                return None
            if not (0 <= prob <= 1):
                print(f"parse_topic_sentiment_distribution_Excel: Value out of [0,1] range at row {row_num} for column B.")
                return None
            if not all(0 <= val <= 1 for val in (c, d, e)):
                print(f"parse_topic_sentiment_distribution_Excel: Values out of [0,1] range at row {row_num} for columns C-E.")
                return None
            if abs((c + d + e) - 1) > 1e-9:
                print(f"parse_topic_sentiment_distribution_Excel: Columns C-E do not sum to 1 at row {row_num}.")
                return None

            # Discard rows with zero proportion in column B
            if prob > 0:
                result[topic] = (prob, (c, d, e))
        
        # Verify that the total probability from column B sums to 1
        total_prob = sum(prob for prob, _ in result.values())
        if abs(total_prob - 1) > 1e-9:
            print("parse_topic_sentiment_distribution_Excel: Sum of column B values does not equal 1.")
            return None

        return result

    except Exception as e:
        print(f"parse_topic_sentiment_distribution_Excel: {e}")
        return None
