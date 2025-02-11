import openpyxl
from pathlib import Path
import matplotlib.pyplot as plt

from topic_chunk_splitter import get_chunks
from chunk_group_allocator import allocate_chunks

def parse_topic_sentiment_distribution_Excel(path: str) -> dict:
    
    file_path = Path(path)
    if not file_path.exists():
        print(f"File does not exist: {path}")
        return {}
    if not file_path.suffix.lower() in ('.xlsx', '.xlsm'):
        print(f"File is not an Excel workbook: {path}")
        return {}
    wb = openpyxl.load_workbook(path)
    
    if "MAIN" not in wb.sheetnames:
        print(f"Error opening workbook or sheet: {e}")
        return {}
    ws = wb["MAIN"]

    result = {}
    row = 4
    while True:
        try:
            cell_a = ws.cell(row=row, column=1).value
        except Exception as e:
            print(f"Error accessing cell A{row}: {e}")
            return {}
        if cell_a is None:
            break
        if not isinstance(cell_a, str):
            print(f"Non-string value in column A at row {row}.")
            return {}

        cell_b = ws.cell(row=row, column=2).value
        cell_c = ws.cell(row=row, column=3).value
        cell_d = ws.cell(row=row, column=4).value
        cell_e = ws.cell(row=row, column=5).value

        if any(not isinstance(x, (int, float)) for x in [cell_b, cell_c, cell_d, cell_e]):
            print(f"Non-numeric value in columns B-E at row {row}.")
            return {}
        if not (0 <= cell_b <= 1):
            print(f"Value out of [0,1] range at row {row} for column B.")
            return {}
        if not (0 <= cell_c <= 1 and 0 <= cell_d <= 1 and 0 <= cell_e <= 1):
            print(f"Values out of [0,1] range at row {row} for columns C-E.")
            return {}
        if abs(cell_c + cell_d + cell_e - 1) > 1e-9:
            print(f"Columns C-E do not sum to 1 at row {row}.")
            return {}

        if cell_b > 0:
            result[cell_a] = (cell_b, (cell_c, cell_d, cell_e))
        
        row += 1

    total_b = sum(item[0] for item in result.values())
    if abs(total_b - 1) > 1e-9:
        print("Sum of column B values does not equal 1.")
        return {}

    return result

def visualize_chunk_allocation(all_chunks):
    # Define sentiment colors
    sentiment_colors = {
        'neg': '#ff9999',  # pastel red
        'neu': '#ffff99',  # yellow
        'pos': '#99ccff'   # pastel blue
    }

    # Sort the reviews by total word count (ascending)
    all_chunks = sorted(all_chunks, key=lambda review: sum(chunk['wc'] for chunk in review))
    
    # Define a custom order for sentiments and sort each review accordingly
    sentiment_order = {'neg': 0, 'neu': 1, 'pos': 2}
    for review in all_chunks:
        review.sort(key=lambda chunk: sentiment_order.get(chunk['sentiment'], 99))
    
    # Create the plot with flipped axes: reviews on the x-axis, word count on the y-axis
    fig, ax = plt.subplots(figsize=(10, 6))
    bar_width = 0.8  # default width for each bar

    # Loop over reviews to plot each as a vertical stacked bar
    for i, review in enumerate(all_chunks):
        bottom = 0  # starting y-coordinate for the stacked segments in this review
        for chunk in review:
            word_count = chunk['wc']
            sentiment = chunk['sentiment']
            
            # Draw one segment (vertical bar) for this chunk
            # We set edgecolor='none' so we can manually add only top and bottom borders
            ax.bar(i, word_count, bottom=bottom, width=bar_width,
                   color=sentiment_colors.get(sentiment, 'gray'),
                   edgecolor='none')
            
            # Compute the left and right edges of the bar for drawing horizontal lines
            left_edge = i - bar_width/2
            right_edge = i + bar_width/2
            
            # Draw a 1px horizontal line at the bottom of the segment
            ax.hlines(y=bottom, xmin=left_edge, xmax=right_edge, colors='black', linewidth=1)
            # Draw a 1px horizontal line at the top of the segment
            ax.hlines(y=bottom + word_count, xmin=left_edge, xmax=right_edge, colors='black', linewidth=1)
            
            # Update bottom for the next segment
            bottom += word_count

    # Set x-ticks and labels
    ax.set_xlabel('Review')
    ax.set_ylabel('Word Count')
    ax.set_title('Stacked Bar Chart of Reviews by Sentiment (Sorted by Total Word Count)')

    plt.tight_layout()
    plt.show()

if __name__ == "__main__":
    
    CHUNK_WC_MIN = 20
    CHUNK_WC_MAX = 50
    TOTAL_WC = 5000
    
    # Get the path relative to the script location
    file_path = Path(__file__).parent / "rulebooks" / "TEMPLATE.xlsx"
    
    ts_distribution_percentage = parse_topic_sentiment_distribution_Excel(file_path)
    if not ts_distribution_percentage:
        print("Error parsing topic-sentiment distribution Excel file.")
    else:
        
        partitioning_error = False
        all_chunks = []
        
        for topic_name, topic_values in ts_distribution_percentage.items():
            if not partitioning_error:
                topic_wc = int(TOTAL_WC * topic_values[0])
                
                for index, sentiment in enumerate(["pos", "neu", "neg"]):
                    topic_sentiment_wc = int(topic_wc * topic_values[1][index])
                    
                    if topic_sentiment_wc == 0:
                        continue
                    chunks = get_chunks(topic_sentiment_wc, CHUNK_WC_MIN, CHUNK_WC_MAX, chunk_count_pref=0.2, dirichlet_a=5.0)
                
                    if not chunks:
                        partitioning_error = True
                        print(f"Error partitioning - topic:'{topic_name}' sentiment:'{sentiment}' - wc:{topic_sentiment_wc}.")
                        break
                    else:
                        for i in chunks:
                            all_chunks.append({'topic': topic_name, 'sentiment': sentiment, 'wc': i})
        
        if not partitioning_error:
            buckets = [
                {'range': (20, 50), 'target_fraction': 0.66},
                {'range': (51, 200), 'target_fraction': 0.01},
                {'range': (201, 300), 'target_fraction': 0.11},
                {'range': (301, 400), 'target_fraction': 0.11},
                {'range': (401, 500), 'target_fraction': 0.11},
            ]
            
            solution = allocate_chunks(all_chunks, buckets, time_limit=20, max_iter=100000)
            
            if not solution:
                print("Error allocating chunks to collections.")
            else:
                collections = [i['chunks'] for i in solution]
                visualize_chunk_allocation(collections)